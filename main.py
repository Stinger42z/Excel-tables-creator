from PyQt6.QtWidgets import (QApplication, QWidget, QToolTip, QTableWidget, QTableWidgetItem,
                             QPushButton, QComboBox, QLineEdit, QLabel, QHeaderView, QVBoxLayout, QHBoxLayout, QMessageBox)
from PyQt6.QtGui import QFont, QIcon, QRegularExpressionValidator
from PyQt6.QtCore import Qt, QRegularExpression
import sys
import pandas as pd
import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import os
from copy import copy

def read_file():
    df = pd.read_excel('./Накладные/Цены.xlsx', header = None, names = ['products', 'prices'])

    places = pd.ExcelFile('./Накладные/Цены.xlsx').sheet_names
    products = df['products'].values.tolist()
    prices = df['prices'].values.tolist()

    return places, products, prices
    
class main_window(QWidget):
    def __init__(self):
        super().__init__()

        self.places, self.products, self.prices = read_file()

        self.initUI()

    # объявление окна
    def initUI(self):
        QToolTip.setFont(QFont('SansSerif', 14))

        # высота панели задач 30 пикселей
        # тогда разрешение экрана 1920х1050
        # self.setGeometry(0,30,1920, 1020)
        self.setGeometry(320, 165, 1280, 720)
        self.setWindowTitle('Печать накладных ФИНАЛЬНАЯ')
        self.setWindowIcon(QIcon('./icon.png'))

        # итоговая сумма
        self.total = QLabel(self)
        self.total.setText(f'Итого: {self.prices[0]:.2f}\t\t')
        bold_font = QFont('SansSerif', 13)
        bold_font.setBold(True)
        self.total.setFont(bold_font)

        # объявление таблицы
        self.table = QTableWidget(self)
        self.table.setColumnCount(7)
        #self.table.resize(1200, 380)
        #self.table.move(40, 100)
        self.table.setRowCount(1)
        self.fill_row()
        #self.table.resizeColumnsToContents()
        header = self.table.horizontalHeader()
        headers = ['№', 'Наименование', 'Единица измерения', 'Количество', 'Цена', 'Стоимость', 'Удаление']
        self.table.setHorizontalHeaderLabels(headers)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)

        # окно выбора получателя
        recipient_label = QLabel(self)
        recipient_label.setText('Кому:')
        recipient_label.setFont(QFont('SansSerif', 12))
        #label.move(40, 18)

        self.recipient = QComboBox(self)
        self.recipient.addItems(self.places)
        self.recipient.setFont(QFont('SansSerif', 11))
        self.recipient.setToolTip('Нажмите, чтобы выбрать получателя')
        self.recipient.setFixedSize(240, 40)
        self.recipient.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)
        self.recipient.currentIndexChanged.connect(lambda: self.recipient_changed())
        #recipient.move(40, 45)

        # окно выбора отправителя
        otprav_label = QLabel(self)
        otprav_label.setText('От кого:')
        otprav_label.setFont(QFont('SansSerif', 12))
        #label.move(40, 18)

        self.otprav = QComboBox(self)
        self.otprav.addItems(['Отправитель 1', 'Отправитель 2'])
        self.otprav.setFont(QFont('SansSerif', 11))
        self.otprav.setToolTip('Нажмите, чтобы выбрать отправителя')
        self.otprav.setFixedSize(240, 40)
        self.otprav.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)
        #self.sender.currentIndexChanged.connect(lambda: self.otprav_changed())

        # кнопка добавления строки
        add_row = QPushButton('+', self)
        add_row.setFont(QFont('SansSerif', 16))
        add_row.setToolTip('Нажми, чтобы добавить еще один товар')
        add_row.setFixedSize(180, 40)
        #add_row.move(40, 500)
        add_row.clicked.connect(lambda: self.f_add_row())

        # кнопка сохранения в pdf/excel
        save = QPushButton('Сохранить в Excel', self)
        save.setFont(QFont('SansSerif', 12))
        save.setToolTip('Нажмите, чтобы сохранить итоговую таблицу в Excel')
        save.setFixedSize(180, 40)
        save.clicked.connect(lambda: self.export())

        # создание слоя для относительного расположения элементов
        self.vertical_layout = QVBoxLayout(self)
        self.vertical_layout.addWidget(otprav_label)
        self.vertical_layout.addWidget(self.otprav)
        self.vertical_layout.addWidget(recipient_label)
        self.vertical_layout.addWidget(self.recipient)
        self.vertical_layout.addWidget(self.table)
        self.vertical_layout.addWidget(self.total, alignment = Qt.AlignmentFlag.AlignRight)
        self.vertical_layout.addWidget(add_row, alignment = Qt.AlignmentFlag.AlignCenter)
        self.vertical_layout.addWidget(save, alignment = Qt.AlignmentFlag.AlignCenter)
        self.vertical_layout.setContentsMargins(20, 20, 20, 25)
        self.vertical_layout.setSpacing(20)


    # заполняет строку дефолтными значениями
    def fill_row(self):
        # номер
        num = QTableWidgetItem(str(self.table.rowCount()))
        num.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(self.table.rowCount()-1, 0, num)

        # наименование
        name = QComboBox(self)
        name.addItems(self.products)
        name.currentIndexChanged.connect(lambda: self.name_changed())
        self.table.setCellWidget(self.table.rowCount()-1, 1, name)

        # еденица измерения
        unit = QLineEdit(self)
        unit.setText('шт.')
        unit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setCellWidget(self.table.rowCount()-1, 2, unit)

        # количество
        amount = QLineEdit(self)
        amount.setAlignment(Qt.AlignmentFlag.AlignCenter)
        amount.setText('1')
        amount.textChanged.connect(lambda: self.amount_changed())
        validator = QRegularExpression(r'[0-9.]+')
        amount.setValidator(QRegularExpressionValidator(validator))
        self.table.setCellWidget(self.table.rowCount()-1, 3, amount)

        # цена
        val = self.prices[0]
        cost = QLineEdit(self)
        cost.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cost.setText('%.2f' % (val))
        cost.textChanged.connect(lambda: self.cost_changed())
        cost.setValidator(QRegularExpressionValidator(validator))
        self.table.setCellWidget(self.table.rowCount()-1, 4, cost)

        # стоимость
        summ = QTableWidgetItem('%.2f' % (1 * float(val)))
        summ.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(self.table.rowCount()-1, 5, summ)

        # удаление
        delete = QPushButton('', self)
        delete.setIcon(QIcon('./bin.png'))
        delete.clicked.connect(lambda: self.remove_row())
        self.table.setCellWidget(self.table.rowCount()-1, 6, delete)

        self.f_total()

    # удаляет строку, и обновляет нумерацию
    def remove_row(self):
        cur_row = self.table.currentRow()

        # удаление строки
        self.table.removeRow(cur_row)

        # обновление нумерации
        row_count = self.table.rowCount()
        for i in range(cur_row, row_count):
            num = QTableWidgetItem(str(i+1))
            num.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(i, 0, num)

        if self.table.rowCount() == 0:
            self.f_add_row()

        self.f_total()
        
    # добавляет строку и заполняет ее дефолтными значениями
    def f_add_row(self):
        count = self.table.rowCount()
        self.table.insertRow(count)
        self.fill_row()

    # при изменении наименования меняется цена
    def name_changed(self):
        # print('Сработала функция name_changed')
        cur_row = self.table.currentRow()

        # изменение цены
        if self.table.cellWidget(cur_row, 1).currentText() in self.products:
            index = self.products.index(self.table.cellWidget(cur_row, 1).currentText())
            val = self.prices[index]
        else:
            val = 0
        
        self.table.cellWidget(cur_row, 4).setText('%.2f' % (val))

        self.f_total()

    # при изменении количества меняется сумма
    def amount_changed(self):
        # print('Сработала функция amount_changed')
        cur_row = self.table.currentRow()
        
        if len(self.table.cellWidget(cur_row, 3).text()) == 0 or len(self.table.cellWidget(cur_row, 4).text()) == 0:

            self.table.item(cur_row, 5).setText('%.2f' % (0))
            return
        
        val = self.table.cellWidget(cur_row, 4).text()
        num = self.table.cellWidget(cur_row, 3).text()
        self.table.item(cur_row, 5).setText('%.2f' % (float(num) * float(val)))

        self.f_total()

    # при изменении цены меняется сумма
    def cost_changed(self):
        # print('Сработала функция cost_changed')
        cur_row = self.table.currentRow()
        
        if len(self.table.cellWidget(cur_row, 4).text()) == 0 or len(self.table.cellWidget(cur_row, 3).text()) == 0:
            self.table.item(cur_row, 5).setText('%.2f' % (0))
            return
        
        num = self.table.cellWidget(cur_row, 3).text()
        val = self.table.cellWidget(cur_row, 4).text()
        self.table.item(cur_row, 5).setText('%.2f' % (float(num) * float(val)))


        self.f_total()

    # функция подсчета итоговой суммы
    def f_total(self):
        summ = 0
        for i in range(self.table.rowCount()):
            summ += float(self.table.item(i, 5).text())
        
        self.total.setText(f'Итого: {summ:.2f}\t\t')

    # функция изменения получателя
    def recipient_changed(self):
        
        page = self.recipient.currentText()
        df = pd.read_excel('./Накладные/Цены.xlsx', header = None, names = ['products', 'prices'], sheet_name = page)
        
        # новые строки заполняются дефольно первыми значениями списков
        self.products = df['products'].values.tolist()
        self.prices = df['prices'].values.tolist()

        # существующие строки остаются такими же, если они есть, если нет, то дефолт
        for i in range(self.table.rowCount()):
            self.table.selectRow(i)
            buffer = self.table.cellWidget(i, 1).currentText()
            self.table.cellWidget(i, 1).clear()
            self.table.cellWidget(i, 1).addItems(self.products)

            if buffer in self.products:
                index = self.products.index(buffer)
                self.table.cellWidget(i, 1).setCurrentIndex(index)

    # функция экспорта данных в Excel
    def export(self):
        # информационное окно
        self.popup_message = QMessageBox(self)
        self.popup_message.setWindowTitle('Внимание')
        self.popup_message.setStandardButtons(QMessageBox.StandardButton.NoButton)
        self.popup_message.setText('Дождитесь завершения работы с Excel')
        self.popup_message.show()
        
        try:
            data = []
            header = ['№', 'Наименование товара', 'Ед. изм.', 'Количество', 'Цена', 'Сумма']

            for row in range(self.table.rowCount()):
                data.append([self.table.item(row, 0).text(),
                            self.table.cellWidget(row, 1).currentText(),
                            self.table.cellWidget(row, 2).text(),
                            self.table.cellWidget(row, 3).text(),
                            self.table.cellWidget(row, 4).text(),
                            self.table.item(row, 5).text()])

            border = Border(left = Side(style = 'thin'), 
                            right = Side(style = 'thin'), 
                            top = Side(style = 'thin'), 
                            bottom = Side(style = 'thin'))
            
            border_line = Border(bottom = Side(style = 'thin'))

            font11 = Font(name = 'Calibri', size = 11, bold = True)
            font12 = Font(name = 'Calibri', size = 12, bold = True)
            font14 = Font(name = 'Calibri', size = 14, bold = True)
            font11bold = Font(name = 'Calibri', size = 11, bold = True)
            font12bold = Font(name = 'Calibri', size = 12, bold = True)
            font14bold = Font(name = 'Calibri', size = 14, bold = True)

            today = date.today()

            month_dict = {'January': '1 Январь',
                        'February': '2 Февраль',
                        'March': '3 Март',
                        'April': '4 Апрель',
                        'May': '5 Май',
                        'June': '6 Июнь',
                        'July': '7 Июль',
                        'August': '8 Август',
                        'September': '9 Сентябрь',
                        'October': '10 Октябрь',
                        'November': '11 Ноябрь',
                        'December': '12 Декабрь'}

            month_rod = {'01': 'января',
                        '02': 'февраля',
                        '03': 'марта',
                        '04': 'апреля',
                        '05': 'мая',
                        '06': 'июня',
                        '07': 'июля',
                        '08': 'августа',
                        '09': 'сентября',
                        '10': 'октября',
                        '11': 'ноября',
                        '12': 'декабря'}

            path = f'./Накладные/{self.recipient.currentText()}/{today.strftime("%Y")}/{month_dict[today.strftime("%B")]}.xlsx'
            
            if os.path.exists(f'./Накладные/{self.recipient.currentText()}/{today.strftime("%Y")}') == False:
                os.makedirs(f'./Накладные/{self.recipient.currentText()}/{today.strftime("%Y")}')

            doc_num = ''
            try:
                file = xl.load_workbook(path)
                for i in range(0, today.day):
                    if i == 0:
                        k = 0
                    else:
                        k = 1
                    last_sheet = (today - timedelta(days = i)).strftime('%d.%m')
                    if last_sheet in file.sheetnames:
                        needed_sheet = file[last_sheet]
                        row = needed_sheet.cell(1, 2).value
                        doc_num = str(int(row.split(' ')[-1]) + k)
                        break
            except:
                file = xl.Workbook()
                doc_num = '1'

            sheet = today.strftime('%d.%m')

            if sheet in file.sheetnames:
                buffer = [x for x in file.sheetnames if x[:5] == sheet]
                if len(buffer) > 1:
                    last_num = len(buffer)
                else:
                    last_num = 1


                doc_num += f'/{len(buffer)}'
                sheet = date.today().strftime('%d.%m') + f' ({int(last_num)})'

            file.create_sheet(sheet)
            current_sheet = file[sheet]

            row1 = f'Накладная № {doc_num}'
            row2 = f'''"{today.strftime('%d')}" {month_rod[today.strftime('%m')]} {today.strftime('%Y')} года'''
            row3 = f'От кого: {self.otprav.currentText()}'
            row4 = f'Кому: {self.recipient.currentText()}'
            rows = [row1, row2, row3, row4]
            rows_font = [font14bold, font12bold, font11bold, font11bold]

            col_last1 = f'Отпустил:{"_" * 15}'
            col_last2 = f'Получил:{"_" * 15}'
            row_last = [col_last1, col_last2]

            block = len(rows)
            table_rows = self.table.rowCount()

            for i in range(1, block + 1):
                current_sheet.cell(i, 2).value = rows[i-1]
                current_sheet.cell(i, 2).font = rows_font[i-1]
                if i < 3:
                    current_sheet.cell(i, 2).alignment = Alignment(horizontal = 'center')

            for i in range(1, len(header) + 1):
                current_sheet.cell(block + 2, i).value = header[i-1]
                current_sheet.cell(block + 2, i).alignment = Alignment(horizontal = 'center')
                current_sheet.cell(block + 2, i).border = border
                current_sheet.cell(block + 2, i).font = font11bold

            for i in range(block + 3, table_rows + block + 3):
                for j in range(1, len(header) + 1):
                    current_sheet.cell(i, j).value = data[i-block-3][j-1]
                    if j == 2:
                        current_sheet.cell(i, j).alignment = Alignment(horizontal = 'left')
                    else:
                        current_sheet.cell(i, j).alignment = Alignment(horizontal = 'center')
                    current_sheet.cell(i, j).border = border

            # строка итого
            current_sheet.cell(table_rows + block + 3, 2).value = 'Итого'
            current_sheet.cell(table_rows + block + 3, 2).alignment = Alignment(horizontal = 'right')
            current_sheet.cell(table_rows + block + 3, 2).font = font11bold
            total = self.total.text().split(' ')[-1].strip()
            current_sheet.cell(table_rows + block + 3, 6).value = total
            current_sheet.cell(table_rows + block + 3, 6).alignment = Alignment(horizontal = 'center')
            current_sheet.cell(table_rows + block + 3, 6).font = font11bold
            for i in range(1, 6 + 1):
                current_sheet.cell(table_rows + block + 3, i).border = border

            current_sheet.cell(table_rows + block + 7, 1).value = row_last[0]
            current_sheet.cell(table_rows + block + 7, 1).font = font11bold
            current_sheet.cell(table_rows + block + 7, 5).value = row_last[1]
            current_sheet.cell(table_rows + block + 7, 5).font = font11bold

            for i in range(1, len(header) + 1):
                current_sheet.cell(table_rows + block + 10, i).border = border_line

            # растяжка столбцов
            col_width = [5, 50, 10, 12, 12, 15]

            for i in range(1, len(col_width) + 1):
                current_sheet.column_dimensions[get_column_letter(i)].width = col_width[i-1]

            # копирую накладную
            # всего строк 14 + len(data)
            start = current_sheet.max_row + 2 # через 1 строку

            for i in range(start, start + 12 + len(data) - 1):
                for j in range(1, len(header) + 1):
                    current_sheet.cell(i, j).value = copy(current_sheet.cell(i-start+1, j).value)
                    current_sheet.cell(i, j).font = copy(current_sheet.cell(i-start+1, j).font)
                    current_sheet.cell(i, j).alignment = copy(current_sheet.cell(i-start+1, j).alignment)
                    current_sheet.cell(i, j).border = copy(current_sheet.cell(i-start+1, j).border)

            # задаю область печати
            current_sheet.sheet_properties.pageSetUpPr.fitToPage = True
            
            file.save(path)
        except Exception as e:
            error = f'{sys.exc_info()[0].__name__}: {e}'

            self.popup_message2 = QMessageBox(self)
            self.popup_message2.setWindowTitle('Ошибка')
            self.popup_message2.setStandardButtons(QMessageBox.StandardButton.Ok)
            self.popup_message2.setText(f'Произошла ошибка.\n{error}')
            self.popup_message2.s
            self.popup_message2.show()

        self.popup_message.hide()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = main_window()
    window.setLayout(window.vertical_layout)
    window.showMaximized()
    app.exec()






